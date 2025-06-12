import pandas as pd
from openpyxl import load_workbook # Using openpyxl to preserve the formatting and fomulas 
from src.input_module import input_results
from src.date_tool import ask_for_date  
from src.workout_logic import give_workout_info

FILE_PATH = '/Users/cole/Documents/OneDrive - State University of New York at New Paltz/Climbing Training Log Summer.xlsx'
def main():
    df = pd.read_excel(FILE_PATH)
    wb = load_workbook(FILE_PATH, data_only=True)  # Load workbook with formulas evaluated
    ws = wb["Plan"]
    print(wb.sheetnames)
    current_date = ask_for_date(df)
    give_workout_info(current_date, df, ws, wb)
    input_results(df, current_date, ws, wb)
#   TODO: generate_report(df)
    
    wb.save(FILE_PATH)
    wb.close()

if __name__ == "__main__":
    main()